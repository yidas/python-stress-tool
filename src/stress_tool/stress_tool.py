import worker_dispatcher as lib_worker_dispatcher
import time, datetime, copy
import openpyxl

default_config = {
    'raw_logs': {
        'fields': {}
    },
}

def start(config: dict, worker_dispatcher: object=None):
    worker_dispatcher = lib_worker_dispatcher if worker_dispatcher is None else worker_dispatcher
    return worker_dispatcher.start(config)

def generate_report(
        config: dict={}, 
        worker_dispatcher: object=None, 
        file_path: str='./tps-report.xlsx',
        display_intervals: bool=True,
        interval: float=0,
        use_processing: bool = False,
        verbose: bool = False,
        debug: bool = False,
        ):

    config =_merge_dicts_recursive(default_config, config)
    worker_dispatcher = lib_worker_dispatcher if worker_dispatcher is None else worker_dispatcher

    # Check the last config
    wd_config = worker_dispatcher.get_last_config()
    if not wd_config.get('worker'):
        return False
    
    # TPS calculation
    tps_data = worker_dispatcher.get_tps(interval=interval, display_intervals=display_intervals, use_processing=use_processing, verbose=verbose, debug=debug)

    # Create a new Workbook
    workbook = openpyxl.Workbook()

    # Sheet for TPS Report
    sheet = workbook.active
    sheet.title = "Report"
    # Taipei timezone
    datetime_gmt = datetime.timezone(datetime.timedelta(hours=time.localtime().tm_gmtoff / 3600))
    # Each row data
    rows_data = []
    rows_data.append([])
    rows_data.append(["TPS", tps_data['tps']])
    rows_data.append(["Started at", datetime.datetime.fromtimestamp(tps_data['started_at'], datetime_gmt).isoformat()])
    rows_data.append(["Ended at", datetime.datetime.fromtimestamp(tps_data['ended_at'], datetime_gmt).isoformat()])
    rows_data.append(["Total Duration", "{:.6f} sec".format(tps_data['duration'])])
    rows_data.append(["Number of Requests", tps_data['count']['total']])
    if wd_config['worker'].get('per_second'):
        rows_data.append(["Concurrency per second", int(wd_config['worker'].get('number') / wd_config['worker'].get('per_second'))])
    else:
        rows_data.append(["Concurrency", wd_config['worker'].get('number')])
    rows_data.append(["Number of Successes", tps_data['count']['success']])
    rows_data.append(["Success Rate", "{:.2f}%".format(tps_data['count']['success'] / tps_data['count']['total'] * 100 if tps_data['count']['total'] > 0 else 0)])
    rows_data.append(["Metrices:"])
    rows_data.append(["Average Execution Time", "{:.6f} sec".format(tps_data['metrics']['execution_time']['avg'])])
    rows_data.append(["Maximum Execution Time", "{:.6f} sec".format(tps_data['metrics']['execution_time']['max'])])
    rows_data.append(["Minimum Execution Time", "{:.6f} sec".format(tps_data['metrics']['execution_time']['min'])])
    rows_data.append(["Success Average Execution Time", "{:.6f} sec".format(tps_data['metrics']['success_execution_time']['avg'])])
    rows_data.append(["Success Maximum Execution Time", "{:.6f} sec".format(tps_data['metrics']['success_execution_time']['max'])])
    rows_data.append(["Success Minimum Execution Time", "{:.6f} sec".format(tps_data['metrics']['success_execution_time']['min'])])
    # Peak TPS
    if tps_data['peak']:
        rows_data.append([])
        rows_data.append(["Peak TPS", tps_data['peak']['tps']])
        rows_data.append(["Peak Started at", datetime.datetime.fromtimestamp(round(tps_data['peak']['started_at'], 3), datetime_gmt).isoformat()])
        rows_data.append(["Peak Ended at", datetime.datetime.fromtimestamp(round(tps_data['peak']['ended_at'], 3), datetime_gmt).isoformat()])
        rows_data.append(["Peak Duration", "{:.6f} sec".format(tps_data['peak']['duration'])])
        rows_data.append(["Peak Number of Requests", tps_data['peak']['count']['total']])
        rows_data.append(["Peak Number of Successes", tps_data['peak']['count']['success']])
        rows_data.append(["Peak Success Rate", "{:.2f}%".format(tps_data['peak']['count']['success'] / tps_data['peak']['count']['total'] * 100)])
        rows_data.append(["Peak Metrices:"])
        rows_data.append(["Average Execution Time", "{:.6f} sec".format(tps_data['peak']['metrics']['execution_time']['avg'])])
        rows_data.append(["Maximum Execution Time", "{:.6f} sec".format(tps_data['peak']['metrics']['execution_time']['max'])])
        rows_data.append(["Minimum Execution Time", "{:.6f} sec".format(tps_data['peak']['metrics']['execution_time']['min'])])
        rows_data.append(["Success Average Execution Time", "{:.6f} sec".format(tps_data['peak']['metrics']['success_execution_time']['avg'])])
        rows_data.append(["Success Maximum Execution Time", "{:.6f} sec".format(tps_data['peak']['metrics']['success_execution_time']['max'])])
        rows_data.append(["Success Minimum Execution Time", "{:.6f} sec".format(tps_data['peak']['metrics']['success_execution_time']['min'])])
    rows_data.append([])
    rows_data.append(["Raw Report", str(tps_data)])
    for row in rows_data:
        sheet.append(row)
    # writer.writerows(rows_data)
    # csv_file.close() 

    # Sheet for Raw Logs
    sheet = workbook.create_sheet(title="Raw Logs")
    # header
    header_row = ["Task ID", "Started at", "Ended at", "Duration (sec)", "Success"]
    # Customized Fields
    customized_fields = config['raw_logs']['fields']
    for key, value in customized_fields.items():
        header_row.append(key)
    sheet.append(header_row)

    # Each row
    for log in worker_dispatcher.get_logs():
        row_data = [log['task_id'], str(log['started_at']), str(log['ended_at']), log['ended_at'] - log['started_at'], worker_dispatcher.result_is_success(log['result'])]
        # Customized Fields
        metadata = log.get('metadata', {})
        for key, value in customized_fields.items():
            if callable(value):
                try:
                    row_data.append(value(metadata))
                except Exception as e:
                    # exit(e)
                    pass
            else:
                row_data.append(str(metadata.get(value, '')))
        sheet.append(row_data)

    # Sheet for Interval
    if tps_data['intervals']:
        sheet = workbook.create_sheet(title="Intervals")
        # header
        sheet.append([
            "TPS", 
            "Started at", 
            "Ended at", 
            "Duration (sec)", 
            "Success", 
            "Total (Done)", 
            "Request",
            "Response",
            "Average Execution Time", 
            "Success Average Execution Time"
        ])
        for row in tps_data['intervals']:
            row_data = [
                row['tps'], 
                str(row['started_at']), 
                str(row['ended_at']), 
                row['duration'], 
                row['count']['success'], 
                row['count']['total'], 
                row['count']['start'], 
                row['count']['end'], 
                row['metrics']['execution_time']['avg'],
                row['metrics']['success_execution_time']['avg']
            ]
            # exit(row_data)
            sheet.append(row_data)

    # Save workbook
    workbook.save(file_path) 
    if verbose: print("TPS Report has been successfully generated at {}".format(file_path))

    return file_path

def _merge_dicts_recursive(default_dict, user_dict):
    merged_dict = copy.deepcopy(default_dict)
    for key, user_value in user_dict.items():
        if key in merged_dict:
            if isinstance(merged_dict[key], dict) and isinstance(user_value, dict):
                merged_dict[key] = _merge_dicts_recursive(merged_dict[key], user_value)
            else:
                merged_dict[key] = user_value
        else:
            merged_dict[key] = user_value
    
    return merged_dict